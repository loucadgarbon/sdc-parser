import csv
import json
import pathlib

from openpyxl import load_workbook
from typer.testing import CliRunner

from sdc_parser.cli import app
from sdc_parser.export_csv import SUMMARY_COLUMNS

FIXTURE = pathlib.Path(__file__).parent / "fixtures" / "sample.tcl"

runner = CliRunner()


def _read_csv(path):
    with open(path, encoding="utf-8-sig", newline="") as fh:
        return list(csv.reader(fh))


def _all_output(result):
    out = result.output
    try:
        out += result.stderr
    except (ValueError, AttributeError):
        pass  # stderr merged into output on this click version
    return out


def _run(*args):
    return runner.invoke(app, [str(a) for a in args])


def test_cli_end_to_end(tmp_path):
    result = _run(FIXTURE, "-o", "report", "--out-dir", tmp_path, "--format", "both")
    assert result.exit_code == 0, _all_output(result)

    summary_rows = _read_csv(tmp_path / "report_summary.csv")
    assert summary_rows[0] == SUMMARY_COLUMNS

    cc = _read_csv(tmp_path / "report_create_clock.csv")
    header, row = cc[0], cc[1]
    assert header[:4] == ["line", "-name", "-period", "arg1"]
    data = dict(zip(header, row))
    assert data["line"] == "6"
    assert data["-name"] == "clk"
    assert data["arg1"] == "[get_ports clk]"
    assert data["active"] == "yes"

    scg = _read_csv(tmp_path / "report_set_clock_groups.csv")
    sdata = dict(zip(scg[0], scg[1]))
    assert sdata["-asynchronous"] == "Y"
    assert sdata["-group"] == "{clk_a clk_b}; {clk_c}"

    wb = load_workbook(tmp_path / "report.xlsx")
    assert wb.sheetnames[0] == "Summary"
    assert "create_clock" in wb.sheetnames

    # terminal summary table is on by default
    assert "Summary" in result.output


def test_cli_params_file(tmp_path):
    params = tmp_path / "params.tcl"
    params.write_text("define MODE scan\n", encoding="utf-8")
    script = tmp_path / "s.tcl"
    script.write_text(
        'if {$MODE == "func"} {\n  puts A\n} else {\n  puts B\n}\n', encoding="utf-8"
    )
    result = _run(
        script, "--params", params, "-o", "pr", "--out-dir", tmp_path, "--format", "csv"
    )
    assert result.exit_code == 0, _all_output(result)
    rows = _read_csv(tmp_path / "pr_puts.csv")
    header = rows[0]
    ai, arg = header.index("active"), header.index("arg1")
    got = {r[arg]: r[ai] for r in rows[1:]}
    assert got == {"A": "no", "B": "yes"}


def test_cli_define_flag(tmp_path):
    script = tmp_path / "s.tcl"
    script.write_text(
        'if {$MODE == "func"} {\n  puts A\n} else {\n  puts B\n}\n'
        "if {$FLAG} { puts F }\n",
        encoding="utf-8",
    )
    result = _run(
        script,
        "-D", "MODE=func", "-D", "FLAG",
        "-o", "d", "--out-dir", tmp_path, "--format", "csv",
    )
    assert result.exit_code == 0, _all_output(result)
    rows = _read_csv(tmp_path / "d_puts.csv")
    header = rows[0]
    ai, arg = header.index("active"), header.index("arg1")
    got = {r[arg]: r[ai] for r in rows[1:]}
    assert got == {"A": "yes", "B": "no", "F": "yes"}


def test_cli_define_overrides_params(tmp_path):
    params = tmp_path / "params.tcl"
    params.write_text("set MODE func\n", encoding="utf-8")
    script = tmp_path / "s.tcl"
    script.write_text('if {$MODE == "scan"} { puts A }\n', encoding="utf-8")
    result = _run(
        script, "--params", params, "-D", "MODE=scan",
        "-o", "o", "--out-dir", tmp_path, "--format", "csv",
    )
    assert result.exit_code == 0
    rows = _read_csv(tmp_path / "o_puts.csv")
    assert dict(zip(rows[0], rows[1]))["active"] == "yes"


def test_cli_bad_define(tmp_path):
    script = tmp_path / "s.tcl"
    script.write_text("puts x\n", encoding="utf-8")
    result = _run(script, "-D", "bad-name=1", "--out-dir", tmp_path)
    assert result.exit_code != 0


def test_cli_filter_active(tmp_path):
    result = _run(
        FIXTURE, "--filter-active", "yes",
        "-o", "fa", "--out-dir", tmp_path, "--format", "csv",
    )
    assert result.exit_code == 0, _all_output(result)
    rows = _read_csv(tmp_path / "fa_puts.csv")
    ai = rows[0].index("active")
    assert rows[1:], "expected surviving puts rows"
    assert all(r[ai] == "yes" for r in rows[1:])
    # summary reflects the same filtered view
    srows = _read_csv(tmp_path / "fa_summary.csv")
    puts_row = dict(
        zip(srows[0], next(r for r in srows[1:] if r[0] == "puts"))
    )
    assert puts_row["active"] == "2 yes"


def test_cli_commands_and_exclude(tmp_path):
    result = _run(
        FIXTURE, "--commands", "set_*",
        "-o", "cm", "--out-dir", tmp_path, "--format", "csv",
    )
    assert result.exit_code == 0
    made = {p.name for p in tmp_path.glob("cm_*.csv")}
    assert "cm_set_false_path.csv" in made
    assert "cm_puts.csv" not in made and "cm_create_clock.csv" not in made

    result = _run(
        FIXTURE, "--exclude", "puts,set",
        "-o", "ex", "--out-dir", tmp_path, "--format", "csv",
    )
    assert result.exit_code == 0
    made = {p.name for p in tmp_path.glob("ex_*.csv")}
    assert "ex_puts.csv" not in made and "ex_set.csv" not in made
    assert "ex_create_clock.csv" in made


def test_cli_fail_on_unknown(tmp_path):
    # fixture has unknown rows inside the proc body -> exit 3
    result = _run(FIXTURE, "--fail-on-unknown", "-o", "fu", "--out-dir", tmp_path)
    assert result.exit_code == 3
    assert "fail-on-unknown" in _all_output(result)

    ok = tmp_path / "ok.tcl"
    ok.write_text("set A 1\nif {$A} { puts P }\n", encoding="utf-8")
    result = _run(ok, "--fail-on-unknown", "-o", "ok", "--out-dir", tmp_path)
    assert result.exit_code == 0


def test_cli_version():
    result = _run("--version")
    assert result.exit_code == 0
    assert "tclscan" in result.output


def test_cli_out_dir_created(tmp_path):
    target = tmp_path / "deep" / "reports"
    result = _run(FIXTURE, "-o", "r", "--out-dir", target, "--format", "csv")
    assert result.exit_code == 0
    assert (target / "r_summary.csv").exists()


def test_cli_quiet_and_no_table(tmp_path):
    result = _run(FIXTURE, "-q", "-o", "q", "--out-dir", tmp_path, "--format", "csv")
    assert result.exit_code == 0
    assert result.output.strip() == ""

    result = _run(
        FIXTURE, "--no-table", "-o", "nt", "--out-dir", tmp_path, "--format", "csv"
    )
    assert result.exit_code == 0
    assert "Summary" not in result.output
    assert "wrote" in result.output


def test_cli_json_and_all_formats(tmp_path):
    result = _run(FIXTURE, "-o", "j", "--out-dir", tmp_path, "--format", "json")
    assert result.exit_code == 0, _all_output(result)
    doc = json.loads((tmp_path / "j.json").read_text(encoding="utf-8"))
    assert doc["schema_version"] == 1
    assert doc["files"] and "sample.tcl" in doc["files"][0]
    assert doc["params"]["MODE"]["value"] == "func"
    cc = doc["commands"]["create_clock"][0]
    assert cc["options"]["-name"] == ["clk"]
    assert cc["positionals"] == ["[get_ports clk]"]
    assert cc["active"] == "yes"
    summary = {s["command"]: s for s in doc["summary"]}
    assert summary["set_false_path"]["active"] == {"yes": 1, "no": 1, "unknown": 0}
    assert any(v["name"] for v in doc["variables"])

    result = _run(FIXTURE, "-o", "a", "--out-dir", tmp_path, "--format", "all")
    assert result.exit_code == 0
    assert (tmp_path / "a.json").exists()
    assert (tmp_path / "a.xlsx").exists()
    assert (tmp_path / "a_summary.csv").exists()


def test_cli_variables_outputs(tmp_path):
    result = _run(FIXTURE, "-o", "v", "--out-dir", tmp_path, "--format", "both")
    assert result.exit_code == 0
    vrows = _read_csv(tmp_path / "v_variables.csv")
    assert vrows[0][0] == "name"
    names = {r[0] for r in vrows[1:]}
    assert {"MODE", "PERIOD"} <= names
    wb = load_workbook(tmp_path / "v.xlsx")
    assert wb.sheetnames[:2] == ["Summary", "Variables"]


def test_cli_xlsx_active_styling(tmp_path):
    result = _run(FIXTURE, "-o", "st", "--out-dir", tmp_path, "--format", "xlsx")
    assert result.exit_code == 0
    wb = load_workbook(tmp_path / "st.xlsx")
    ws = wb["set_false_path"]
    header = [c.value for c in ws[1]]
    ai = header.index("active") + 1
    by_active = {ws.cell(row=r, column=ai).value: r for r in (2, 3)}
    no_row = by_active["no"]
    assert ws.cell(row=no_row, column=1).font.italic
    yes_row = by_active["yes"]
    assert not ws.cell(row=yes_row, column=1).font.italic
    # unknown active cell gets the yellow fill
    ws_set = wb["set"]
    sheader = [c.value for c in ws_set[1]]
    sai = sheader.index("active") + 1
    unknown_rows = [
        r for r in range(2, ws_set.max_row + 1)
        if ws_set.cell(row=r, column=sai).value == "unknown"
    ]
    assert unknown_rows
    assert ws_set.cell(row=unknown_rows[0], column=sai).fill.fgColor.rgb.endswith(
        "FFEB9C"
    )


def test_cli_source_following_e2e(tmp_path):
    (tmp_path / "sub.tcl").write_text("create_clock -name c2 p2\n", encoding="utf-8")
    main = tmp_path / "main.tcl"
    main.write_text(
        "create_clock -name c1 p1\nsource sub.tcl\n", encoding="utf-8"
    )
    result = _run(main, "-o", "src", "--out-dir", tmp_path, "--format", "csv")
    assert result.exit_code == 0
    rows = _read_csv(tmp_path / "src_create_clock.csv")
    assert rows[0][0] == "file"
    assert len(rows) == 3

    result = _run(
        main, "--no-follow-source", "-o", "nsrc", "--out-dir", tmp_path,
        "--format", "csv",
    )
    assert result.exit_code == 0
    rows = _read_csv(tmp_path / "nsrc_create_clock.csv")
    assert rows[0][0] == "line"
    assert len(rows) == 2


def test_cli_unroll_e2e(tmp_path):
    script = tmp_path / "u.tcl"
    script.write_text(
        "foreach clk {a b} {\n  create_clock -name $clk\n}\n", encoding="utf-8"
    )
    result = _run(
        script, "--unroll", "-o", "u", "--out-dir", tmp_path, "--format", "csv"
    )
    assert result.exit_code == 0
    rows = _read_csv(tmp_path / "u_create_clock.csv")
    header = rows[0]
    ni = header.index("-name")
    ei = header.index("arguments_expanded")
    assert len(rows) == 3
    assert {rows[1][ei], rows[2][ei]} == {"-name a", "-name b"}
    assert rows[1][ni] == "$clk"  # original text keeps the variable


def test_cli_diff_mode(tmp_path):
    script = tmp_path / "d.tcl"
    script.write_text(
        'if {$MODE == "func"} {\n  set_false_path -from f\n} else {\n'
        "  set_false_path -from g\n}\n",
        encoding="utf-8",
    )
    result = _run(
        script,
        "-D", "MODE=func",
        "--diff", script, "--diff-define", "MODE=scan",
        "-o", "df", "--out-dir", tmp_path, "--format", "csv",
    )
    assert result.exit_code == 0, _all_output(result)
    assert "diff vs" in result.output
    rows = _read_csv(tmp_path / "df_diff.csv")
    header = rows[0]
    ci = header.index("change")
    fi = header.index("changed_fields")
    cmdi = header.index("command")
    changed = [r for r in rows[1:] if r[ci] == "changed"]
    assert changed
    sfp_changed = [r for r in changed if r[cmdi] == "set_false_path"]
    assert len(sfp_changed) == 2  # both branches flip active
    assert all("active" in r[fi] for r in sfp_changed)

    # identical sides -> no differences
    result = _run(
        script, "-D", "MODE=func", "--diff", script,
        "-o", "same", "--out-dir", tmp_path, "--format", "csv",
    )
    assert result.exit_code == 0
    assert "no differences" in result.output
    assert _read_csv(tmp_path / "same_diff.csv")[1:] == []


def test_cli_missing_file(tmp_path):
    result = _run(tmp_path / "nope.tcl")
    assert result.exit_code == 1
    assert "error" in _all_output(result)


def test_cli_parse_error_and_tolerant(tmp_path):
    bad = tmp_path / "bad.tcl"
    bad.write_text("set ok 1\nset x {unclosed\n", encoding="utf-8")
    result = _run(bad, "-o", "bad", "--out-dir", tmp_path, "--format", "csv")
    assert result.exit_code == 1

    result = _run(
        bad, "-o", "bad", "--out-dir", tmp_path, "--format", "csv", "--tolerant"
    )
    assert result.exit_code == 0
    assert "warning" in _all_output(result)
    assert (tmp_path / "bad_set.csv").exists()
