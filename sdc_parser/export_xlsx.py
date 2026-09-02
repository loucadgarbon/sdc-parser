"""Excel export: one workbook with a Summary sheet followed by one sheet per
command, each with per-argument columns."""

from __future__ import annotations

import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .export_csv import SUMMARY_COLUMNS, VARIABLES_COLUMNS

_WRAP_COLUMNS = {"raw", "conditions"}
_MAX_WIDTH = 60

_INACTIVE_FONT = Font(italic=True, color="808080")
_UNKNOWN_FILL = PatternFill("solid", fgColor="FFEB9C")
_UNKNOWN_FONT = Font(color="9C6500")
_UNRESOLVED_FONT = Font(color="C00000", bold=True)
_SHEET_BAD_CHARS = re.compile(r"[\[\]:*?/\\]")


def _sheet_title(name: str, used: set[str]) -> str:
    title = _SHEET_BAD_CHARS.sub("_", name)[:31] or "_"
    candidate = title
    n = 2
    while candidate.lower() in used:
        suffix = f"~{n}"
        candidate = title[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def _fill_sheet(ws, columns: list[str], rows: list[list], style_active: bool = False):
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wrap_idx = {i for i, c in enumerate(columns) if c in _WRAP_COLUMNS}
    active_idx = columns.index("active") if style_active and "active" in columns else None
    widths = [len(str(c)) for c in columns]
    for row in rows:
        ws.append(row)
        for i, value in enumerate(row):
            text = str(value)
            longest = max((len(part) for part in text.split("\n")), default=0)
            widths[i] = max(widths[i], longest)
        for i in wrap_idx:
            ws.cell(row=ws.max_row, column=i + 1).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
        if active_idx is not None:
            state = row[active_idx]
            if state == "no":
                for cell in ws[ws.max_row]:
                    cell.font = _INACTIVE_FONT
            elif state == "unknown":
                cell = ws.cell(row=ws.max_row, column=active_idx + 1)
                cell.fill = _UNKNOWN_FILL
                cell.font = _UNKNOWN_FONT
    for i, width in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = min(width + 2, _MAX_WIDTH)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_xlsx(
    path: str,
    tables: dict,
    summary_rows: list[dict],
    variables_rows: list[dict] | None = None,
    diff_table: tuple[list[str], list[list]] | None = None,
):
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    _fill_sheet(
        summary,
        SUMMARY_COLUMNS,
        [[row[c] for c in SUMMARY_COLUMNS] for row in summary_rows],
    )
    used = {"summary"}
    if variables_rows:
        ws = wb.create_sheet(_sheet_title("Variables", used))
        _fill_sheet(
            ws,
            VARIABLES_COLUMNS,
            [[row[c] for c in VARIABLES_COLUMNS] for row in variables_rows],
        )
        for offset, row in enumerate(variables_rows):
            if row["unresolved_uses"]:
                ws.cell(row=offset + 2, column=1).font = _UNRESOLVED_FONT
    if diff_table is not None:
        ws = wb.create_sheet(_sheet_title("Diff", used))
        _fill_sheet(ws, diff_table[0], diff_table[1])
    for command, table in tables.items():
        ws = wb.create_sheet(_sheet_title(command, used))
        _fill_sheet(ws, table["columns"], table["rows"], style_active=True)
    wb.save(path)
